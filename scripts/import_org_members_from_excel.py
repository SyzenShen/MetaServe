#!/usr/bin/env python3
"""
批量导入组织成员脚本
使用方法：
  python3 scripts/import_org_members_from_excel.py \
    --file "/home/mosserver/software/Download_system_project/第二期生物信息培训班学员名单.xlsx" \
    --org "第二期生物信息培训班学员" \
    --password "123456Aa"

说明：
  - 读取 Excel 第一列或包含“邮箱”关键字的列，去重后批量创建用户
  - 若用户已存在则跳过创建，仅加入组织
  - 组织不存在时以当前执行用户作为 owner 创建组织
  - 成员默认角色为 member
"""

import argparse
import os
import sys
import django

# 将项目根目录加入 Python 路径，确保可导入 file_project
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'file_project.settings')
django.setup()

import re
from typing import List
from django.contrib.auth import get_user_model
from django.db import transaction
from authentication.models import Organization, Membership

try:
    import openpyxl
except Exception:
    openpyxl = None


def extract_emails_from_excel(path: str) -> List[str]:
    if openpyxl is None:
        raise RuntimeError('缺少 openpyxl 依赖，请先安装：pip install openpyxl')
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active
    emails: List[str] = []
    email_regex = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
    # 找到可能的邮箱列（包含“邮箱”或 Email 的表头），否则默认第一列
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    email_col_idx = 1
    for idx, val in enumerate(header_row, start=1):
        s = str(val or '').strip().lower()
        if s and ('邮箱' in s or 'email' in s):
            email_col_idx = idx
            break

    for row in sheet.iter_rows(min_row=2, values_only=True):
        cell = row[email_col_idx - 1] if len(row) >= email_col_idx else None
        v = str(cell or '').strip()
        if v and email_regex.match(v):
            emails.append(v.lower())
    # 去重
    return sorted(set(emails))


def main():
    parser = argparse.ArgumentParser(description='批量导入组织成员（Excel）')
    parser.add_argument('--file', required=True, help='Excel 文件路径')
    parser.add_argument('--org', required=True, help='组织名称')
    parser.add_argument('--password', required=True, help='统一初始密码')
    args = parser.parse_args()

    path = os.path.abspath(args.file)
    if not os.path.exists(path):
        print({'ok': False, 'message': f'文件不存在: {path}'})
        sys.exit(1)

    emails = extract_emails_from_excel(path)
    if not emails:
        print({'ok': False, 'message': '未在 Excel 中提取到邮箱'})
        sys.exit(1)

    User = get_user_model()

    # 组织：若不存在则创建（使用系统第一个超级用户或第一个用户作为 owner）
    org = Organization.objects.filter(name=args.org).first()
    if not org:
        owner = User.objects.filter(is_superuser=True).first() or User.objects.first()
        if not owner:
            print({'ok': False, 'message': '系统中不存在可作为 owner 的用户，请先创建一个用户'})
            sys.exit(1)
        org = Organization.objects.create(name=args.org, owner=owner)
        Membership.objects.create(organization=org, user=owner, role='owner')

    created_count = 0
    added_count = 0
    with transaction.atomic():
        for email in emails:
            user = User.objects.filter(email=email).first()
            if not user:
                user = User.objects.create_user(email=email, password=args.password)
                created_count += 1
            # 加入组织（默认 member）
            _, created_rel = Membership.objects.update_or_create(
                organization=org,
                user=user,
                defaults={'role': 'member'}
            )
            if created_rel:
                added_count += 1

    print({'ok': True, 'organization': org.name, 'extracted': len(emails), 'users_created': created_count, 'members_added': added_count})


if __name__ == '__main__':
    main()